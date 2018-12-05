# -*- coding: utf-8 -*-
import  os, sys, shutil
import struct
import h5py
import numpy as np
import base64, zlib
import scipy.io
import openpyxl

MatDataIO_version = '2018-06-07'

def Mk_Matlab7_3_text_matrix( id_list ):
    # Put text from list into 16-bit integer matrix for matlab compatible text matrices in hdf5 style matlab 7.3 format
    OutList = []
    maxlen = max( map( len, id_list))
    if maxlen == 0:
        maxlen = 1
        """
        empty strings does not work for the present hdf5 implementation """
        
    for ID in id_list:
#        print('ID', ID)
        ID_mapped = list(map( ord, ID.ljust( maxlen)))

#        print('ID_mapped',ID_mapped )
        OutList.append(ID_mapped)
    TextMatrix = np.mat( OutList, dtype=np.uint16 )
    return TextMatrix

def Rd_Matlab7_3_text_matrix( TextMatrix ):
    id_list = [] 
    for row in range(0, TextMatrix.shape[0]):
        LettersID = map( chr, np.squeeze(np.asarray(TextMatrix[row]))  )
        id_list.append( ''.join(LettersID).strip() )
    return id_list

def Save_folded_v73_mat( out_filename, ImageStack, IDs, SecondIDs ):
    Stackfile = Matlab7_3_file( out_filename )
    Stackfile.WriteMatrix('IMAGES', ImageStack )
    Stackfile.WriteIds( 'VARID1', IDs )
    Stackfile.WriteIds( 'VARID2', SecondIDs )
    Stackfile.close()
    
#def Save_folded_v5_mat( out_filename, ImageStack, IDs, SecondIDs ):
#    print 'Saving :', out_filename 
#    OutData = {}
#    OutData['IMAGES'] = ImageStack
#    maxlen = max( map( len, IDs))
#    OutData['VARID1'] = []
#    for VarID in IDs: 
#        OutData['VARID1'].append( VarID.ljust( maxlen ) ) # did not work to rectify text in Matlab
#        
###    OutData['VARID2'] = secondary_varid02
#
#    # Text not working in matlab -v5 so far ??
#    scipy.io.savemat( out_filename, OutData )


def some_data():
    data_out = np.zeros( (6 , 999) )
    data_out[ 0, :] = np.arange(1,1000)
    data_out[ 1, :] = np.arange(1,1000)
    data_out[ 2, :] = np.arange(1,1000)*2
    data_out[ 3, :] = np.arange(1,1000)*2
    data_out[ 4, :] = np.arange(1,1000)*3
    data_out[ 5, :] = np.arange(1,1000)*3
    return data_out

def some_text():
    text_out = []
    text_out.append( 'abc' )
    text_out.append( 'abcdefgh' )
    text_out.append( '12345678' )
    return text_out
    
def is_monotonic(x):
    if np.all(np.isfinite(x)):    
        dx = np.diff(x)
        return np.all(dx < 0) or np.all(dx > 0)
    else:
        return False
    
def is_same(x):    
    dx = np.diff(x)
    return np.all(dx == 0)

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
        
def is_integer(s):
    try:
        i = int(s)
        if i == float(s):
            return True
        else:
            return False
    except ValueError:
        return False
    
    
def is_HDF_Mat73_file( FileName):
    trace = True
    f = open(FileName, 'rb')
    try:
        Filetype_ID = f.read(19)
        data0 = struct.unpack('<ccccccccccccccccccc', Filetype_ID )
        data = b''.join(data0)
        is_HDF = (data == b'MATLAB 7.0 MAT-file')
        if trace:
            if is_HDF:
                print('This is a HDF_mat7.3 file')
            else:
                print('Filetype_ID:', data)
    finally:
        f.close()
    return is_HDF
    
def get_ObsIds(D):
    existing_ObsIDs = []
    for key in sorted(D.keys()):
        if (len(key) == 6):
            if (key[0:5] == 'OBSID') and is_integer(key[5]):
                existing_ObsIDs.append(key)
    return existing_ObsIDs
    
def get_VarIds(D):
    existing_VarIDs = []
    for key in sorted(D.keys()):
        if (len(key) == 6):
            if (key[0:5] == 'VARID') and is_integer(key[5]):
                existing_VarIDs.append(key)
    return existing_VarIDs

def insert_primary_ObsID(D):
    existing_ObsIDs = get_ObsIds(D)               
#    print('existing_ObsIDs',existing_ObsIDs)
    for i in range(len(existing_ObsIDs), 0, -1):
        D['OBSID'+str(i+1)] = D['OBSID'+str(i)]
#        print(i, '->', i+1)
    D['OBSID1'] = np.arange(D['DATA'].shape[0])
    
def remove_invalid_Simca_keys(D):
    Simca_invalid_keys = []
    Simca_valid_keys = ['DATA']
    for key_num in range(9):
        Simca_valid_keys.append('OBSID'+str(key_num+1))
        Simca_valid_keys.append('VARID'+str(key_num+1))
        
    for key in D:
        if not key in Simca_valid_keys:
            Simca_invalid_keys.append(key)
    for key in Simca_invalid_keys:
        D.pop(key)     
    return D
        
 

def split_Simca_mat_observations(D, split_size):
    """D: dictionary containg Simca mat data
       split_size: number of bytes before split
    """
    bytes_per_observation = (D['DATA'].nbytes+10)/D['DATA'].shape[0]
    obs_per_split_unit = int(split_size/bytes_per_observation)
    num_of_splits = int(np.ceil(D['DATA'].shape[0]/obs_per_split_unit))
    print('num_of_splits',num_of_splits)
    current_ObsIDs = get_ObsIds(D)
    current_VarIDs = get_VarIds(D)
    splitted_dicts = []
    for split in range(num_of_splits-1):
        D_splitted = {}
        D_splitted['DATA'] = D['DATA'][split*obs_per_split_unit:(split+1)*obs_per_split_unit,:]
        for VarID in current_VarIDs:
            D_splitted[VarID] = D[VarID]  
        for ObsID in current_ObsIDs:
            if isinstance(D[ObsID][0], np.unicode_) or \
               isinstance(D[ObsID], list):
                D_splitted[ObsID] = D[ObsID][split*obs_per_split_unit:(split+1)*obs_per_split_unit]
            else:
                if D[ObsID].ndim > 1:
                    D_splitted[ObsID] = D[ObsID][split*obs_per_split_unit:(split+1)*obs_per_split_unit,:]
                else:
                    D_splitted[ObsID] = D[ObsID][split*obs_per_split_unit:(split+1)*obs_per_split_unit]
                    
        splitted_dicts.append(D_splitted)

    last_split = num_of_splits-1
    D_last_split = {}
    D_last_split['DATA'] = D['DATA'][last_split*obs_per_split_unit:,:]
    for VarID in current_VarIDs:
        D_last_split[VarID] = D[VarID] 
    for ObsID in current_ObsIDs:
        if isinstance(D[ObsID][0], np.unicode_) or \
           isinstance(D[ObsID], list):
            D_last_split[ObsID] = D[ObsID][last_split*obs_per_split_unit:]
        else:
            if D[ObsID].ndim > 1:
                D_last_split[ObsID] = D[ObsID][last_split*obs_per_split_unit:,:]
            else:
                D_last_split[ObsID] = D[ObsID][last_split*obs_per_split_unit:] 
    splitted_dicts.append(D_last_split)
    
    return splitted_dicts
 
   
def check_input_dict( D):
    print()
    for key in sorted(D.keys()):
        print( 'key: ', key, end=' ')
        if isinstance(D[key][0], np.unicode_):
            print(D[key].shape, 'np.unicode_')               
        elif isinstance(D[key], list):
            print(len(D[key]), 'list')
        else:
            print(D[key].shape, end=' ')
            if D[key].ndim == 3:
                print(type(D[key][0,0,0]) )
            elif D[key].ndim == 2: 
                print(type(D[key][0,0]) )
            else:
                print(type(D[key][0]) )


def write_keys(PrintName, M, additional_obsids=[], additional_varids=[] ):
    print()
    print(PrintName, 'file keys:')

    for key in sorted(M.keys()):
        if isinstance(M[key], list):
            print(key, 'type:', type(M[key]), type(M[key][0]), flush=True )
            print( '    ', key, 'len('+str(len(M[key]))+')', sys.getsizeof(M[key])/1000000.0, 'Mb,', 'list', end=' ')
            if (key[0:5] == 'OBSID') or (key[0:5] == 'VARID') or key in additional_obsids or key in additional_varids:
                if len(M[key]) == len(set(M[key])):
                    print(' - unique IDs', end=' ')

        else:            
            if M[key].ndim ==0:
                print(key, 'type:', type(M[key]), 'M[key].ndim =', M[key].ndim, flush=True )
                if isinstance(M[key], str):
                    print( '    ', key, len(M[key]), 'str', M[key])
                
                elif isinstance(M[key], np.str_):
                    print( '    ', key, len(M[key]), 'np.str_', end=' ' )
                    print()
                elif isinstance(M[key], np.str_):
                    print( '    ', key, len(M[key]), 'np.str_', end=' ' )
                    print()
                elif isinstance(M[key], np.unicode_):
                    print( '    ', key, M[key].shape, M[key].nbytes/1000000.0, 'Mb,', 'np.unicode_', end=' ' )
                    if (key[0:5] == 'OBSID') or (key[0:5] == 'VARID') or key in additional_obsids or key in additional_varids:
                        if len(M[key]) == len(set(M[key])):
                            print(' - unique IDs', end=' ')
                else:
                    print( '    ', key, M[key].shape, M[key].nbytes/1000000.0, 'Mb,', 'array', end=' ')
            else:
                print(key, 'type:', type(M[key]), type(M[key][0]), flush=True )
        
         
                if isinstance(M[key], str):
                    print( '    ', key, len(M[key]), 'str', M[key])
                
                elif isinstance(M[key][0], np.str_):
                    print( '    ', key, len(M[key]), 'np.str_', end=' ' )
                    print()
                elif isinstance(M[key][0], np.str_):
                    print( '    ', key, len(M[key]), 'np.str_', end=' ' )
                    print()
                elif isinstance(M[key][0], np.unicode_):
                    print( '    ', key, M[key].shape, M[key].nbytes/1000000.0, 'Mb,', 'np.unicode_', end=' ' )
                    if (key[0:5] == 'OBSID') or (key[0:5] == 'VARID') or key in additional_obsids or key in additional_varids:
                        if len(M[key]) == len(set(M[key])):
                            print(' - unique IDs', end=' ')
        
                else:
                    print( '    ', key, M[key].shape, M[key].nbytes/1000000.0, 'Mb,', 'array', end=' ')
                    if (key[0:5] == 'OBSID') or (key[0:5] == 'VARID') or key in additional_obsids or key in additional_varids:
                        if max(M[key].shape) == 1:
                            print(' - single ID', end=' ')
                        else:
                            if len(list(np.squeeze(np.asarray(M[key])))) == len(set(list(np.squeeze(np.asarray(M[key]))))):
                                print(' - unique IDs', end=' ')
                            if is_monotonic(np.squeeze(M[key])):
                                print(' - are monotonic', end=' ')
                            elif is_same(np.squeeze(M[key])):
                                print(' - are same', end=' ')                    
        print('')
    

def adjust_specified_1D_arrays(input_matrix0, VariableName):
    # Simca matlab format needs columns for OBSID and VARID
    if input_matrix0.ndim == 1:
        input_matrix = np.mat(input_matrix0)
        if ((VariableName[0:5] == 'OBSID') or \
            (VariableName[0:5] == 'VARID') or \
            (VariableName[0:4] == 'Z_ID') or \
            (VariableName[0:5] == 'SMPID')) and (input_matrix.shape[0] == 1):
                input_matrix = input_matrix.T
    else:
        input_matrix = input_matrix0
    return input_matrix



def Insert_table(ImageTable, sheet, row_cnt, is_pretty=True):
    trace = False
    TwoDecStyle = '0.00'
    IntStyle = '##0'
    SciStyle = '0.00E+000'        
    for input_row in ImageTable:
        row = [] 
        for entry in input_row:
            if is_number(entry) and isinstance(entry, np.int64):
                row.append(float(entry)) #openpyxl gave an error for numeric integers
            else:
                row.append(entry)
        if trace:
            if (row_cnt % 100) == 0:
                print('Appending - row_cnt:', row_cnt)        
#            print(row)
        sheet.append(row)
        if is_pretty:
            for col in range(0,len(row)):
                if trace:
                    print('row_cnt, col, row[col]', row_cnt, col, row[col])
                if not is_number(row[col]):
                    pass
                elif is_integer(row[col]):                
                    sheet.cell(row=row_cnt+1, column=col+1).number_format = IntStyle
                elif abs(float(row[col])) >= 0.01:
                    sheet.cell(row=row_cnt+1, column=col+1).number_format = TwoDecStyle
                else:
                    sheet.cell(row=row_cnt+1, column=col+1).number_format = SciStyle
                
        row_cnt += 1
    return sheet
    

def Save_XLSX_table(FileName, sheet_name, header_rows, Data):
    trace = False

    if isinstance(Data, np.ndarray):
        DataTable = list(Data)     
    elif isinstance(Data, list):
        DataTable = Data
    else:
        print('Save_XLSX_table input data type not yet handled')
      

    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    if trace:
        print()
        print( 'XLSX sheet names ', sheet.title )
    sheet.title = sheet_name
     
#     for table_entry in range(len(sheet_names)):               
#        sheet = wb.create_sheet(title=sheet_names[table_entry])   
    if isinstance(header_rows, list):
        if all(isinstance(elem, list) for elem in header_rows):
            for row_cnt, header_row in enumerate(header_rows):
        #        print()
        #        print(row_cnt)
        #        print(header_row)
                sheet.append(header_row)
        else: # single row header
            sheet.append(header_rows)
            row_cnt = 1
    else:
        print('The entered table header was not a list')
            
    sheet = Insert_table(DataTable, sheet, row_cnt)

    if trace:
        print()
        print( 'XLSX sheet names ', wb.sheetnames )
    file_suffix = '.xlsx'
    current_file_suffix = os.path.splitext(FileName)[1].lower()
    if current_file_suffix == file_suffix:      
        wb.save(FileName)
    else:
        wb.save(FileName+file_suffix)
    
    

    

class Matlab4_file:
    def __init__(self, Output_file_name):
        self.outfile01 = open( Output_file_name, 'wb')

    def WriteMatrix( self, VariableName0, data0 ):
        # WriteMatrix( VariableName, data )
        # Write a numpy matrix <data> on the filehandle <outfile01> with the variable name <VariableName>
        # in Matlab V4 binary file format as obtained from the 'save filename matrixname .... -v4' 
        # command in Matlab
        
        if data0.ndim > 2:
            raise ValueError( '{mat_name} can only be 1D or 2D for the "v4" matlab format.'.format(mat_name = VariableName0 ))
        else:
            data = np.asarray(adjust_specified_1D_arrays(data0, VariableName0))
            VariableName0 = VariableName0+chr(0) #ascii strings need to be zero terminated 
            VariableName = VariableName0.encode('ascii') # will not work with unicode
#            print('Matlab4_file WriteMatrix data.shape', VariableName0, data.shape, type(data) )
#            print 'VariableName ', VariableName[0:5]
            
            MOPT_doubles = 0
            mrows = data.shape[0]
            ncols = data.shape[1]
            imagf = 0
            namelen = len(VariableName)
                        
            self.outfile01.write(struct.pack('iiiii', MOPT_doubles, mrows, ncols, imagf, namelen ))
            for ch in VariableName:
#                print('ch', ch, type(ch))
                self.outfile01.write(struct.pack('b', ch ))
            for element in data.T.flat: 
                self.outfile01.write(struct.pack('d', element))

     


    def WriteIds( self, VariableName0, IdList ):
        # WriteIds( VariableName, IdList )
        # Write a list of strings <IdList> for a matlab variable on the filehandle <outfile01> 
        # with the variable name <VariableName> in Matlab V4 binary file format as obtained from the
        # 'save filename matrixname .... -v4' command in Matlab

        VariableName0 = VariableName0+chr(0)
        VariableName = VariableName0.encode('ascii', 'replace') # will not work with unicode 
        maxlen = max( map( len, IdList))
        IdList2 = []
        for word in IdList:   #fill up w blanks and convert to numeric codes
#            print( 'WriteIds word', word, word.ljust( maxlen) )
            word2 = word.ljust( maxlen).encode('ascii', 'replace') # stop extended chars to Matlab
#            print( 'WriteIds word2', word2 )# , word2.ljust( maxlen) )
            list_from_map = list(map( int, word2)) # Transform ascii to a number for each letter
            IdList2.append(list_from_map)
        VariableIds = np.mat(IdList2)

        MOPT_text = 1
        mrows = VariableIds.shape[0]
        ncols = VariableIds.shape[1]
        imagf = 0
        namelen = len(VariableName)

        self.outfile01.write(struct.pack('iiiii', MOPT_text, mrows, ncols, imagf, namelen ))
        for ch in VariableName:
#            print('WriteIds VariableName ch', ch, type(ch))
            self.outfile01.write(struct.pack('b', ch ))
        for kod in VariableIds.T.flat: 
            self.outfile01.write(struct.pack('d', float(kod)))
            
    def close(self):
        self.outfile01.close()


# How to embed an image in code:
# import base64, zlib
# template_path=r'/home/anamj/ProgDev/Python2.6/HDF5_as_matlab_file/2012-06-20 Make and use empty template/Empty_mat_template.hdf5'
# data = open( template_path, 'rb').read()
# print base64.encodestring(zlib.compress(data))


class Matlab7_3_file:

    def __init__(self, Output_file_name):
        # 'Empty_mat_template.hdf5' previously created by Matlab, with all matrices deleted in a second step, 
        # is used as an empty file with a matlab signature, to be recognised as a valid Matlab file
        # This is the version with the Template file compressed and embedded in the code

        init_blk = """eJztlc1OwkAQx7cUsUFJRC8Np7mYeEBSEBR70QolkFggSuJRK5RIUsQUTDwZrr6FT+Fz+Ag+gm+A
                u+2WhorR6MGv+SXNzH9nZ7PdzmwNrXWoHcBORgFDa212e7aVhqZtjroDp69Cs3RSq6eh5FjmyOrA
                4FKFitMD7coBZRuyWTW3peZ3Iadkc1AtVwowbF9YfROUjFKADMyH+ERqBkG+lTv61RLxVNwVkkSi
                ZIWIzI948QmnmJrVfv696FmB6zF3zriu6lrTHeda8ueF9uHn33D7/Mn3+W20jnSd2UkIP/4U9aw0
                Px354wi0EVlvANdJ2p9Mh/ssmL/o9RKPL9DKYa4gRtwRkciuTsIer6l9kl8i+cm0j2XP8v7n5Te1
                Et0JX8/VG7EgHiVBna5yj81df7htLBOF6wSNSMRw/zunbdscDtn4Gn1iwbZJZ3B9blvMK37onN7i
                uN4oC3RVf1+P4peW+7ewe4qdY/ieGr+TV5Rn9et6QhAEQRAEQRAEQRAEQX4iLz68Z1k="""

        Template = zlib.decompress(base64.decodestring( init_blk.encode('ascii')))
        try:
            Template_file = open( Output_file_name, 'wb')
            Template_file.write(Template)
            Template_file.close()
        except IOError as e:
            print( 'Error in Create_matlab7_3_file at copying the empty template to the new file ')
            errno, strerror = e.args
            print( "I/O error({0}): {1}".format(errno,strerror))
            
        self.hdf5_file_handle = h5py.File(Output_file_name, 'a' )
    

##    def __init__(self, Output_file_name):
##        # 'Empty_mat_template.hdf5' previously created by Matlab, with all matrices deleted in a second step, 
##        # is used as an empty file with a matlab signature, to be recognised as a valid Matlab file
##        appPath = os.path.abspath(os.path.dirname(os.path.join(sys.argv[0])))
##        Template_file_name = 'Empty_mat_template.hdf5'
##        FullPath_template_hdf5_file_name = os.path.join(appPath, Template_file_name)
##        try:
##            shutil.copy(FullPath_template_hdf5_file_name, Output_file_name)
##            self.hdf5_file_handle = h5py.File(Output_file_name, 'a' )
##        except Exception, e:
##            print 'Error in Create_matlab7_3_file at copying the empty template to the new file ', e



    def WriteMatrix( self, VariableName, input_matrix0 ): 
        # WriteMatrix( VariableName, input_matrix )
        # Write a numpy matrix <input_matrix> on the hdf5_file_handle with the variable name <VariableName>
        # in Matlab V7.3 binary file format as obtained from the 'save filename matrixname .... -v7.3' 
        # command in Matlab
##        self.hdf5_file_handle.create_dataset(VariableName, data=input_matrix.T, compression=4, chunks=input_matrix.T.shape )
        if input_matrix0.ndim == 1:
#            input_matrix = np.mat(np.asarray(input_matrix0, dtype=np.float))
            input_matrix = np.mat(input_matrix0)
            if ((VariableName[0:5] == 'OBSID') or \
                (VariableName[0:5] == 'VARID') or \
                (VariableName[0:4] == 'Z_ID') or \
                (VariableName[0:5] == 'SMPID')) and (input_matrix.shape[0] == 1):
                    input_matrix = input_matrix.T
        else:
#            input_matrix = np.asarray(input_matrix0, dtype=np.float)
            input_matrix = input_matrix0

        self.hdf5_file_handle.create_dataset(VariableName, data=input_matrix.T, compression=4 ) #  default chunking
##        self.hdf5_file_handle.create_dataset( VariableName, data=input_matrix.T ) # No compression, default chunking (works above 4 Gb)
        dset1str = '/'+VariableName
        dset1 =self.hdf5_file_handle[dset1str]
        if isinstance(input_matrix[0,0], np.float64):
            dset1.attrs[ 'MATLAB_class' ] = b'double'
        elif isinstance(input_matrix[0,0], np.float32):
            dset1.attrs[ 'MATLAB_class' ] = b'single'
        elif isinstance(input_matrix[0,0], np.int64):
            dset1.attrs[ 'MATLAB_class' ] = b'int64'
        elif isinstance(input_matrix[0,0], np.int32):    
            dset1.attrs[ 'MATLAB_class' ] = b'int32'
        elif isinstance(input_matrix[0,0], np.uint32):    
            dset1.attrs[ 'MATLAB_class' ] = b'uint32'
        elif isinstance(input_matrix[0,0], np.uint8):    
            dset1.attrs[ 'MATLAB_class' ] = b'uint8'
        elif isinstance(input_matrix[0,0], np.bool_):    
            dset1.attrs[ 'MATLAB_class' ] = b'logical'
        else:
            print('Warning: unrecognized type in WriteMatrix v73 format:', type(input_matrix[0,0]), 'MATLAB_class not set')
    
    def WriteIds( self, VariableName, input_list ):
        # WriteIds( VariableName, input_list )
        # Write a list of strings <input_list> for a matlab variable on the hdf5_file_handle
        # with the variable name <VariableName> in Matlab V7.3 binary file format as obtained from the
        # 'save filename matrixname .... -v7.3' command in Matlab
        ID_mat = Mk_Matlab7_3_text_matrix( input_list )
        dset2str = '/'+VariableName
        dset2 = self.hdf5_file_handle.create_dataset(dset2str, data=ID_mat.T, compression=4, chunks=ID_mat.T.shape ) 
        dset2.attrs[ 'MATLAB_class' ] = 'char'
        dset2.attrs[ 'MATLAB_int_decode' ] = 2
        
        
    def close(self):
        self.hdf5_file_handle.close()
        

def WriteDataChunk( OutFileName, chunk, version='v4' ):
    
    is_debug = False
    if version == 'v4':    
        Mat_out = Matlab4_file( OutFileName )
        if is_debug:
            print( 'Saving file in matlab -v4 format: ', OutFileName )
    elif version in ['v7.3', 'v73']:
        Mat_out = Matlab7_3_file( OutFileName )
        if is_debug:
            print( 'Saving file in matlab -v7.3 format: ', OutFileName )
    else:
        raise ValueError( '{current_inp} is wrong, only "v4" and "v7.3" are implemented so far'.format(current_inp = version ))

    try:
        for key in sorted(chunk):
            if is_debug:
                print( 'mat-key: ', key )
            if isinstance(chunk[key], str):
                if is_debug:
                    print('    case str', key, chunk[key])
                Mat_out.WriteIds( key, [chunk[key]])
            
            elif isinstance(chunk[key], list):
                if is_debug:
                    print('    case list')
                Mat_out.WriteIds( key, chunk[key])   
            
            elif  chunk[key].ndim == 0:
                if isinstance(chunk[key], np.str_):
                    if is_debug:
                        print('    case np.str_ ndim=0')
                    as_list = []
                    for index in range(0, len(chunk[key])):
                        if chunk[key][index]:
                            as_list.append( chunk[key][index] )
                        else:
                            as_list.append(' ')
                    Mat_out.WriteIds( key, as_list)
                    
                elif isinstance(chunk[key], np.unicode_):
                    if is_debug:
                        print('    case np.unicode_ ndim=0')
                    as_list = []
                    for index in range(0, len(chunk[key])):
                        if chunk[key][index]:
                            as_list.append( chunk[key][index] )
                        else:
                            as_list.append(' ')
                    Mat_out.WriteIds( key, as_list)
                else:    
                    if is_debug:
                        print('    case scalar')
                    Mat_out.WriteMatrix( key, np.ones((1,1))*chunk[key])
                
            else:
                if isinstance(chunk[key][0], np.str_):
                    if is_debug:
                        print('    case np.str_ array')
                    as_list = []
                    for index in range(0, len(chunk[key])):
                        as_list.append( chunk[key][index] )
                    Mat_out.WriteIds( key, as_list)
                    
                elif isinstance(chunk[key][0], np.unicode_):
                    if is_debug:
                        print('    case np.unicode_ array')
                    as_list = []
                    for index in range(0, len(chunk[key])):
                        as_list.append( chunk[key][index] )   
                    Mat_out.WriteIds( key, as_list)
                    
                else:
                    if is_debug:
                        print('    case array')
                    Mat_out.WriteMatrix( key, chunk[key])        
    finally:           
        Mat_out.close()
 

def get_size_info(list_or_array):
    if isinstance(list_or_array, list):
        size_info = len(list_or_array)                    
    else:
        size_info = list_or_array.shape
    return size_info            
      

def WriteData(FileName, D, version='v4', is_version_suffix=True, approx_max_file_size = 1.95E9):    
    trace = False
    generic_filename = os.path.splitext(FileName)[0]
    if is_version_suffix:
        version_suffix = '_'+version
    else:
        version_suffix = ''
    if (version == 'v4'):            
        #Done 'DATA' may not always be prsent in a IMAGE data set 
        if not ('DATA' in D):
            OutFileName = generic_filename +version_suffix+'.mat'
            WriteDataChunk( OutFileName, D, version )
        else: # Able to split into many files if data table is too big for Simca input
            current_data = np.asarray(np.atleast_2d(D['DATA']))
            if isinstance(D['OBSID1'], list):
                rows_are_ok_no_transpose = (current_data.shape[0] == len(D['OBSID1']))
                rows_ok_to_transpose = (current_data.shape[1] == len(D['OBSID1']))
            else:
                if trace:
                    print("D['OBSID1'].shape", D['OBSID1'].shape)
                rows_are_ok_no_transpose = (current_data.shape[0] == max(D['OBSID1'].shape))
                rows_ok_to_transpose = (current_data.shape[1] == max(D['OBSID1'].shape))
            
            if isinstance(D['VARID1'], list):
                cols_are_ok_no_transpose = (current_data.shape[1] == len(D['VARID1']))
                cols_ok_to_transpose = (current_data.shape[0] == len(D['VARID1']))
            else:
                if trace:
                    print("D['VARID1'].shape", D['VARID1'].shape)
                cols_are_ok_no_transpose = (current_data.shape[1] == max(D['VARID1'].shape))
                cols_ok_to_transpose = (current_data.shape[0] == max(D['VARID1'].shape))
            
                
            if rows_ok_to_transpose and cols_ok_to_transpose:
                D['DATA'] = current_data.T
            elif rows_are_ok_no_transpose and cols_are_ok_no_transpose:
                D['DATA'] = current_data
            else:                
                print('*** DATA shape is:', D['DATA'].shape, 'while OBSID1 shape is:', get_size_info(D['OBSID1']), 'and VARID1 shape is:', get_size_info(D['VARID1']) )

            bytes_per_observation = (D['DATA'].nbytes+10)/D['DATA'].shape[0]
            max_num_of_observations_in_one_chunk = int(approx_max_file_size/bytes_per_observation)
                
            if (D['DATA'].shape[0] > max_num_of_observations_in_one_chunk):
                if trace:
                    print()
                    print('bytes_per_observation',bytes_per_observation)
                    print('observations_per', round(approx_max_file_size/1.0E09, 3), 'Gb :', max_num_of_observations_in_one_chunk)
                    
                splitted_dicts = split_Simca_mat_observations(D, approx_max_file_size)        
                chunk_cnt = 1
                for chunk in splitted_dicts:
                    if trace:
                        write_keys('chunk'+str(chunk_cnt), chunk )
                    OutFileName = generic_filename +version_suffix+'_chunk_'+str(chunk_cnt)+'.mat'
                    WriteDataChunk( OutFileName, chunk, version )
                    chunk_cnt += 1
            
            else:
                OutFileName = generic_filename +version_suffix+'.mat'
                WriteDataChunk( OutFileName, D, version )
                
    elif version in ['v7.3', 'v73']:
        OutFileName = generic_filename +version_suffix+'.mat'
        WriteDataChunk( OutFileName, D, version )
    else:
        raise ValueError( '{current_inp} is wrong, only "v4" and "v7.3" are implemented so far'.format(current_inp = version ))
        
        

#def ReadData_hdf5_object(FileName, version='v7.3'):
#    trace = False
#    if trace:
#        print('Enter ReadData_hdf5_object'  )
#    if version == 'v7.3':
#        hdf5_obj = h5py.File( FileName, 'r' )        
#    else:
#        print("Version ID was not recognized, needs to be: 'v4', 'v5' or 'v7.3'")
#        hdf5_obj = None
#    return hdf5_obj       
    
    
def ReadData(FileName, version='v7.3'):
    Data = {}
    trace = False
    if trace:
        print('Enter ReadData version', version  )
    if version == 'v7.3':
        hdf5_obj = h5py.File( FileName, 'r' )
        for key in hdf5_obj.keys():
            if trace:
                print( 'hdf5_obj key=', key, ' matlab_class:', hdf5_obj[key].attrs['MATLAB_class'], hdf5_obj[key].shape)
#            if not isinstance(hdf5_obj[key], np.ndarray):
#                Data[key] = np.ones((1,1))*hdf5_obj[key][:]
            if (hdf5_obj[key].attrs['MATLAB_class'] == 'double') or (hdf5_obj[key].attrs['MATLAB_class'] == b'double'):
                if (hdf5_obj[key][:].ndim == 2) and ((hdf5_obj[key][:].shape[0] == 1) or (hdf5_obj[key][:].shape[1] == 1)):
                    Data[key] = np.squeeze(np.asarray(hdf5_obj[key][:] )) #remove redundant dimension
                else:
                    Data[key] = hdf5_obj[key][:].T
                if trace:
                    print('    size=', Data[key].shape )
                    
            elif (hdf5_obj[key].attrs['MATLAB_class'] == 'logical') or (hdf5_obj[key].attrs['MATLAB_class'] == b'logical'):
                if (hdf5_obj[key][:].ndim == 2) and ((hdf5_obj[key][:].shape[0] == 1) or (hdf5_obj[key][:].shape[1] == 1)):
                    Data[key] = np.squeeze(np.asarray(hdf5_obj[key][:], dtype=np.bool )) #remove redundant dimension
                else:
                    Data[key] = np.asarray(hdf5_obj[key][:].T, dtype=np.bool)
                    
            elif (hdf5_obj[key].attrs['MATLAB_class'] == 'int32') or (hdf5_obj[key].attrs['MATLAB_class'] == b'int32') or \
                 (hdf5_obj[key].attrs['MATLAB_class'] == 'int64') or (hdf5_obj[key].attrs['MATLAB_class'] == b'int64'):

                if hdf5_obj[key].ndim == 0:
                    Data[key] = np.ones((1,1))*hdf5_obj[key][:]                
                elif (hdf5_obj[key][:].ndim == 2) and ((hdf5_obj[key][:].shape[0] == 1) or (hdf5_obj[key][:].shape[1] == 1)):
                    Data[key] = np.squeeze(np.asarray(hdf5_obj[key][:] )).astype(np.float64) #remove redundant dimension
                else:
                    Data[key] = hdf5_obj[key][:].T.astype(np.float64)
                if trace:
                    print('    size=', Data[key].shape )
                
            elif (hdf5_obj[key].attrs['MATLAB_class'] == 'char') or (hdf5_obj[key].attrs['MATLAB_class'] == b'char'):
                mat_chars = hdf5_obj[key][:].T
                matlist = Rd_Matlab7_3_text_matrix(mat_chars)
                Data[key] = matlist                
            else:
                print('The hdf5 MATLAB_class attribute', hdf5_obj[key].attrs['MATLAB_class'], ' was not recognized.')
        hdf5_obj.close()
            
    elif 'v4' or 'v5':
        Data = scipy.io.loadmat(FileName)
        for key in Data.keys():
            if not isinstance(Data[key], list):
                if (Data[key].ndim == 2) and ((Data[key].shape[0] == 1) or (Data[key].shape[1] == 1)):
                    Data[key] = np.squeeze(np.asarray( Data[key] ))                    
                
    else:
        print("Version ID was not recognized, needs to be: 'v4', 'v5' or 'v7.3'")
        
    return Data
           

def WriteData_as_xlsx(FileName, D):
    is_debug = False
    sheet_counter = 0    
    wb = openpyxl.Workbook()
    try:
        for key in sorted(D):
            sheet = wb.create_sheet(key)
#            sheet.title = key
            if is_debug:
                print( 'xlsx key: ', key )
            if isinstance(D[key], str):
                if is_debug:
                    print('1')
                Insert_table([[D[key]]], sheet, 0)
            elif isinstance(D[key][0], np.str_):
                if is_debug:
                    print('2')
                as_list = []
                for index in range(0, len(D[key])):
                    as_list.append( [D[key][index]] )
                Insert_table(as_list, sheet, 0)
                
            elif isinstance(D[key][0], np.unicode_):
                if is_debug:
                    print('3')
                as_list = []
                for index in range(0, len(D[key])):
                    as_list.append( [D[key][index]] )
                Insert_table(as_list, sheet, 0)
                
            elif isinstance(D[key], list):
                if is_debug:
                    print('4')
                list_of_rows = []
                for row in D[key]:
                    list_of_rows.append([row])    
                Insert_table(list_of_rows, sheet, 0)
            elif isinstance(D[key], np.ndarray):
                if is_debug:
                    print('5')
                if D[key].ndim == 2:
                    data_in_list = list(D[key])
                elif D[key].ndim == 1:
                    data_in_list = list(np.expand_dims(D[key], axis=0))                    
                else:
                    print('The xlsx file cannot handle more than 2D arrays')                
                
                Insert_table(data_in_list, sheet, 0)
                
                if is_debug:   
                    if D[key].ndim > 1:
                        print('case 5 2D', 'data_in_list', type(data_in_list), type(data_in_list[0]), type(data_in_list[0][0]) )
                    else:
                        print('case 5 1D', 'data_in_list', type(data_in_list), type(data_in_list[0]))
                    print()
                
            else:
                if is_debug:
                    print('6') 
                print('Key=', key, 'A data type was not handled well on its way to the xlsx file')
            sheet_counter += 1
            
    finally:           
        file_suffix = '.xlsx'
        if os.path.splitext(FileName)[1].lower == file_suffix:        
            wb.save(FileName)
        else:
            wb.save(FileName+file_suffix)

    

if __name__ == '__main__':
    print( os.path.basename(sys.argv[0]), end='')
    print( ' Version: ', MatDataIO_version )

    Output_file_name = 'Matlab7_3_test01.mat'
    # Test writing with basic primitives
    matfile = Matlab7_3_file( Output_file_name )
    matfile.WriteMatrix( 'DATA', some_data() )
    matfile.WriteIds( 'VARID1', some_text() )
    matfile.close()
    
    Output_file_name = 'Matlab4_test01.mat'
    # Test writing with basic primitives
    matfile = Matlab4_file( Output_file_name )
    matfile.WriteMatrix( 'DATA', some_data() )
    matfile.WriteIds( 'VARID1', some_text() )
    matfile.close()
    
    Output_file_name = 'Matlab4_test02.mat'
    # Test writing with WriteData
    merged_dict = {'DATA':some_data(),'VARID1':some_text() }
    WriteData( Output_file_name, merged_dict )
    
    Output_file_name = 'Matlab7_3_test02.mat'
    # Test writing with WriteData
    merged_dict = {'DATA':some_data(),'VARID1':some_text() }
    WriteData( Output_file_name, merged_dict, version='v7.3' )
    
    print( 'Tests of '+ os.path.basename(sys.argv[0]) +' finished execution')