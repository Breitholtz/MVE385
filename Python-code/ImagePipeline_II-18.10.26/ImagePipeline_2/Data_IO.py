# -*- coding: utf-8 -*-
"""
Created on Sat Jul 19 18:29:57 2014

@author: mats_j
"""
import os, sys, time
import datetime
import platform
import openpyxl
import numpy as np

#import xlwt
#import HTML
#import pyexiv2
#import numpy as np
#from PIL import Image
#import matplotlib.pyplot as plt
#import MVA_package.markup as markup
#import MVA_package.statistical_functions as stat

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def is_integer(s):
    try:
        i = int(s)
        if i == float(s):
            return True
        else:
            return False
    except ValueError:
        return False

#def nanhistogram(a, bins=10, range=None, weights=None, density=None):
#    not_NaNs = np.isfinite(a)
#    # proportion_of_finites = np.count_nonzero(a[not_NaNs]) / (1.0*a.shape[0]*a.shape[1])
#    # print 'proportion_of_finites ', proportion_of_finites
#    hist, bin_edges = np.histogram(a[not_NaNs], bins, range, weights, density)
#    return hist, bin_edges
#
#
def make_dir_if_absent(d):
    if not os.path.exists(d):
        os.makedirs(d)


def getConfigFile(NameOfSoftware):
    pure_software_name = os.path.basename(os.path.splitext(NameOfSoftware)[0])
    ConfigFileName = pure_software_name+'.config'
    if platform.system() == 'Darwin':#added
        os_settings_file_location = os.path.join(os.path.expanduser('~'),'.'+ pure_software_name )
    if platform.system() == 'Linux':
        os_settings_file_location = os.path.join(os.path.expanduser('~'),'.'+ pure_software_name )
    elif platform.system() == 'Windows':
        os_settings_file_location = os.path.join('C:\\ProgramData', pure_software_name)
    config_file_path = os.path.join(os_settings_file_location, ConfigFileName)
    return config_file_path


#def insert_exif_info(FileName, exif_info ):
#    metadata = pyexiv2.ImageMetadata(FileName)
#    metadata.read()
#    Exif_Key1 = 'Exif.Photo.UserComment'
#    metadata[Exif_Key1] = pyexiv2.ExifTag(Exif_Key1, exif_info['UserComment'])
#    xmp_key1 = 'Xmp.dc.title'
#    metadata[xmp_key1] = exif_info['UserComment']
#    metadata.write()



def Insert_table(ImageTable, sheet, row_cnt):
    trace = False
    TwoDecStyle = '0.00'
    IntStyle = '##0'
    SciStyle = '0.00E+000'
    for input_row in ImageTable:
        row = []
        for entry in input_row:
            if is_number(entry) and isinstance(entry, np.int64):
                row.append(float(entry)) #openpyxl gave an error for numeric integers
            else:
                row.append(entry)
        if trace:
            print('Appending - row_cnt:', row_cnt)
            print(row)
        sheet.append(row)
        for col in range(0,len(row)):
            if trace:
                print('row_cnt, col, row[col]', row_cnt, col, row[col])
            if not is_number(row[col]):
                pass
            elif is_integer(row[col]):
                sheet.cell(row=row_cnt+1, column=col+1).number_format = IntStyle
            elif abs(float(row[col])) >= 0.01:
                sheet.cell(row=row_cnt+1, column=col+1).number_format = TwoDecStyle
            else:
                sheet.cell(row=row_cnt+1, column=col+1).number_format = SciStyle

        row_cnt += 1
    return sheet


def Save_XLSX_table(FileName, sheet_name, header_rows, ImageTable ):
    trace = False
#    wb = openpyxl.Workbook(encoding='utf-8') # explicit encoding ccaused an error
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    if trace:
        print()
        print( 'XLSX sheet names ', sheet.title )
    sheet.title = sheet_name

#     for table_entry in range(len(sheet_names)):
#        sheet = wb.create_sheet(title=sheet_names[table_entry])
    row_cnt = 0
    for header_row in header_rows:
        sheet.append(header_row)
        row_cnt += 1
    sheet = Insert_table(ImageTable, sheet, row_cnt)

    if trace:
        print()
        print( 'XLSX sheet names ', wb.sheetnames )
    wb.save(FileName)

def datetime_stamp():
    stamp = '%4d-%02d-%02d_%02d.%02d.%02d' % time.localtime()[:6]
    return stamp

def timestamped_filename( base_filename ):
    filename = base_filename+datetime_stamp()
    return filename
#
#
#
#def Make_html_page( ListOfOrigImageFileNames, ListOfFiles, ListOfImageSizes, SubDir, ImagesPerSection, component_names_in_RGB_order ):
#
#    is_color = (len(component_names_in_RGB_order) > 0)
#    is_ImageSizes = (len(ListOfImageSizes) > 0)
#    title = "Imagepipeline images"
#    header = "Place for menu links"
#    max_colors_file_name = 'max_colors.png'
#    page = markup.page( mode='xml' )
#
#    page.init( title=title, header=header, encoding='UTF-8' )
#    page.body( style="color: white; background-color: black;", link="#66ffff")
#    #page(escape=False)
#    page.br( )
#
#    cnt = 0
#    max_file = len(ListOfFiles)
#    shown_file_names = []
#    page.br( )
#    while cnt < max_file:
#        if cnt < max_file:
#            #if SubDir != '':
#            src_local_raw_img_name = os.path.join( SubDir, os.path.basename(ListOfFiles[cnt])) # already encoded at making of list
#            max_color_img_name = os.path.join( SubDir, max_colors_file_name )
#            #else: src_local_raw_img_name = os.path.join( SubDir, os.path.basename(ListOfFiles[cnt]))
#            orig_raw_img_name = ListOfOrigImageFileNames[cnt].encode('ascii', 'xmlcharrefreplace')
#            if is_ImageSizes:
#                h = ListOfImageSizes[cnt][0]
#                w = ListOfImageSizes[cnt][1]
#                page.img( alt=src_local_raw_img_name, src=src_local_raw_img_name, height=h, width=w )
#            else:
#                page.img( alt=src_local_raw_img_name, src=src_local_raw_img_name )
#            shown_file_names.append(str(cnt+1)+', ' + orig_raw_img_name)
#        cnt += 1
#        if (cnt % ImagesPerSection) == 0:
#            if is_color:
#                page.br( )
#                page.a('RGB:')
#                for c_name in component_names_in_RGB_order:
#                    page.a(c_name)
#                page.img( alt=max_color_img_name, src=max_color_img_name )
#            for file_name_to_show in shown_file_names:
#                page.br( )
#                page
#                page.small(file_name_to_show)
#            page.br( )
#            shown_file_names = []
#
#
#    if len(shown_file_names) > 0:
#        if is_color:
#            page.br( )
#            page.a('RGB:')
#            for c_name in component_names_in_RGB_order:
#                page.a(c_name)
#            page.img( alt=max_color_img_name, src=max_color_img_name )
#    for file_name_to_show in shown_file_names: # print the remaning file names
#        page.br( )
#        page.small(file_name_to_show)
#    page.br( )
#    html_text_out = str(page)
#    return html_text_out
#
#
#
#def Save_rounded_html_table(FileName, header_row3, ImageTable3 ):
#
#    web_rows_header = []
#    web_rows = []
#    web_excluded_cols = [9]
#    num_of_columns = len(header_row3)
#    try:
#        for list_col in range(0, num_of_columns):
#            if not (list_col in web_excluded_cols):
#                web_rows_header.append(header_row3[list_col].replace('_', ' ') )
#
#        for line in ImageTable3:
#            web_row = []
#            for list_col in range(0, num_of_columns):
#                if not (list_col in web_excluded_cols):
#                    if not is_number(line[list_col]):
#                            web_row.append( line[list_col] )
#                    elif is_integer(line[list_col]):
#                        heltal = int(line[list_col])
#                        if heltal == 0:
#                            web_row.append( '0' )
#                        else:
#                            web_row.append( int(line[list_col]) )
#                    elif abs(float(line[list_col])) >= 0.01:
#                        web_row.append( "%.3f" % float(line[list_col]) )
#                    else:
#                        web_row.append( '%.2E' % float(line[list_col]) )
#            web_rows.append(web_row)
#
#    finally:
#        htmlcode_1 = HTML.table(web_rows, col_width=(3, 70, 5), header_row=web_rows_header)
#        f_1 = open( FileName, 'w' )
#        f_1.write(htmlcode_1)
#        f_1.close()
#
#
##def normalize_into_0_1_range( Layer0 ):
##    # Layer is 2D matrix
##    Layer = (Layer0 - np.nanmin(Layer0) )/(np.nanmax(Layer0)  - np.nanmin(Layer0))
##    return Layer
#
#
#def GetBWimageMaps(ImageStack, ImgMin, ImgMax):
#    NumOfImages = ImageStack.shape[2]
#    ImageMapInt = np.zeros( (ImageStack.shape[0], ImageStack.shape[1], NumOfImages), dtype=np.uint8 )
#    for ImgNum in range(0, NumOfImages):
#        current_Layer = ImageStack[ :, :, ImgNum]
#        Layer = stat.normalize_into_0_1_range_min_max( current_Layer, ImgMin[ImgNum], ImgMax[ImgNum]  )
#        ImageMapInt[:,:, ImgNum ] = np.uint8( 255*Layer )
#    return ImageMapInt
#
#
#def GetHistograms(ImageStack, ImgMin, ImgMax):
#    NumOfBins = 256
#    NumOfImages = ImageStack.shape[2]
#    ImageHistogram = np.zeros( (NumOfBins, NumOfImages), dtype=np.int64 )
#    for ImgNum in range(0, NumOfImages):
#        current_Layer = ImageStack[ :, :, ImgNum]
#        hist, bin_edges = nanhistogram(current_Layer, bins=NumOfBins, range=(ImgMin[ImgNum], ImgMax[ImgNum]))
#        ImageHistogram[:, ImgNum ] = hist
#    return ImageHistogram
#
#
#def make_histogram(hist, title, limits):
#    hist_x_scale = np.linspace(limits[0], limits[1], num=hist.shape[0] )
#    y_origin = np.zeros_like( hist )
#    fig0 = plt.figure()
#    ax0 = fig0.add_subplot(1, 1, 1)
#    ax0.vlines( hist_x_scale, y_origin, hist, color='b', linestyles='solid')
#    ax0.axvline(limits[0], color='g')
#    ax0.axvline(limits[1], color='r')
#    ax0.set_xlabel('magnitude')
#    ax0.set_ylabel('counts')
#    ax0.set_title(title)
#    return fig0
#
##def save_histogram_image(hist, overall_min, overall_max, FileName):
##
##    current_fig = make_histogram(hist, (overall_min, overall_max))
##    current_fig.savefig(FileName)
##    plt.close( current_fig )
#
#
#def Export_image_histograms(ResultSet_Names, Image_matrices, ImgMin, ImgMax, Out_FileDir):
#    ListOfOrigImageFileNames = []
#    ListOfBWHistoFileNames = []
#    ListOfImageSizes = []
#    is_png = False
#    SubDir = 'histo_BW'
#    ImgDir = os.path.join(Out_FileDir, SubDir)
#    make_dir_if_absent(ImgDir)
#    for R_Name in ResultSet_Names:
#        Result = Image_matrices[ R_Name ]
#        tracing_info = Result['Name']
#        ImageHistogram = GetHistograms(Result['ImageStack'], ImgMin, ImgMax)
#        NumOfComponents = ImageHistogram.shape[1]
#        for ImgIx in range(0, NumOfComponents):
#            hist_out = ImageHistogram[:,ImgIx]
#            Clean_ascii_name = Result['Name'].encode('ascii','replace').replace('?','_') +'_'+ str(ImgIx+1)
#            if is_png:
#                hist_out_name = os.path.join(ImgDir, Clean_ascii_name + '_histo_BW.png')
#            else:
#                hist_out_name = os.path.join(ImgDir, Clean_ascii_name + '_histo_BW.svg')
#            visible_name = Result['Name'] +' '+ Result['VARID1'][ImgIx]
#            current_fig = make_histogram(hist_out, visible_name, (ImgMin[ImgIx], ImgMax[ImgIx]))
#            current_fig.savefig(hist_out_name)
#            ListOfOrigImageFileNames.append(visible_name)
#            dpi = current_fig.get_dpi() /2 # div 2 to get 3x4 inch from 6x8 matplotlib standard
#            ListOfImageSizes.append( ( int(dpi*current_fig.get_figheight()), int(dpi*current_fig.get_figwidth()) ) )
#            plt.close( current_fig )
#
#            ListOfBWHistoFileNames.append( hist_out_name )
##            ListOfOrigImageFileNames.append(Result['Name'] +' '+ Result['VARID1'][ImgIx])
#            if is_png:
#                insert_exif_info(hist_out_name, {'UserComment':tracing_info +' '+ Result['VARID1'][ImgIx]})
#
#    return ListOfOrigImageFileNames, ListOfBWHistoFileNames, ListOfImageSizes, SubDir, NumOfComponents
#
#
#
#
#def Export_BW_images(ResultSet_Names, Image_matrices, ImgMin, ImgMax, Out_FileDir):
#    ListOfOrigImageFileNames = []
#    ListOfBWImageFileNames = []
#    ListOfImageSizes = []
#    SubDir = 'raw_BW'
#    ImgDir = os.path.join(Out_FileDir, SubDir)
#    make_dir_if_absent(ImgDir)
#    for R_Name in ResultSet_Names:
#        Result = Image_matrices[ R_Name ]
#        tracing_info = Result['Name']
#        ImageMaps = GetBWimageMaps(Result['ImageStack'], ImgMin, ImgMax)
#        NumOfComponents = ImageMaps.shape[2]
#        for ImgIx in range(0, NumOfComponents):
#            ListOfImageSizes.append( (ImageMaps[:,:,ImgIx].shape[0], ImageMaps[:,:,ImgIx].shape[1]) )
#            img_out = Image.fromarray( ImageMaps[:,:,ImgIx] )
#            Clean_ascii_name = Result['Name'].encode('ascii','replace').replace('?','_') +'_'+ str(ImgIx+1)
#            img_out_name = os.path.join(ImgDir, Clean_ascii_name + '_raw_BW.png')
#            img_out.save( img_out_name )
#            ListOfBWImageFileNames.append( img_out_name )
#            ListOfOrigImageFileNames.append(Result['Name'] +' '+ Result['VARID1'][ImgIx])
#            insert_exif_info(img_out_name, {'UserComment':tracing_info +' '+ Result['VARID1'][ImgIx]})
#
#    return ListOfOrigImageFileNames, ListOfBWImageFileNames, ListOfImageSizes, SubDir, NumOfComponents
#
#
#



class DataRepository:
    """ The DataRepository class is meant as a handler of log-files and output results in a common
        directory structure using dates, software component names and numbers as directory names.
        It is intended as a common results repository for more than one software, if desired. """

    def __init__(self):
        self.IsRepositoryReady = False
        self.CurrentLocation = 'DataRepositoryLocation.txt'
        if platform.system() == 'Darwin':
            self.os_settings_file_location = os.path.join(os.path.expanduser('~'),'.data_repository' )
        if platform.system() == 'Linux':
            self.os_settings_file_location = os.path.join(os.path.expanduser('~'),'.data_repository' )
        elif platform.system() == 'Windows':
            self.os_settings_file_location = 'C:\\ProgramData\\MVA_data_repository'
        self.settings_file_path =  os.path.join(self.os_settings_file_location, self.CurrentLocation)
        # print('DataRepository self.settings_file_path ', self.settings_file_path)
        if os.path.exists(self.settings_file_path):
            self.IsRepositoryReady = True


    def set_location(self, location_of_repository):
        make_dir_if_absent(os.path.dirname(self.settings_file_path))
        settings_file = open(self.settings_file_path, 'w')
        settings_file.write(location_of_repository)
        self.IsRepositoryReady = True
        settings_file.close()


    def get_location(self):
        if self.IsRepositoryReady:
            settings_file = open(self.settings_file_path, 'r')
            Repository_directory = settings_file.readline()
            settings_file.close()
            return Repository_directory
        else:
            return None

    def IsDefined(self):
        return self.IsRepositoryReady


    def get_pos_integer(self, s):
        try:
            i = int(s)
            if i == float(s):
                return int(s)
            else:
                return 0
        except ValueError:
            return 0

    def get_next_numeric_directory_index(self, dir0):
        dir_content = os.listdir(dir0)
        numlist = map(self.get_pos_integer, dir_content )
        if numlist:
            dir_index = max(numlist) + 1
        else:
            dir_index = 1
        return dir_index


    def get_new_entry_in_data_repository(self, RepositoryId, repository_dir=''):
        """
        RepositoryId: becomes directory level name below date
        repository_dir: if empty, use pre-set directory, otherwise use input directory"""

        if not (repository_dir or self.IsRepositoryReady):
            return None
        elif repository_dir:
            pass
        elif self.IsRepositoryReady:
            repository_dir = self.get_location()

        date_stamp = '%4d-%02d-%02d' % time.localtime()[:3]
        todays_dir = os.path.join( repository_dir, date_stamp )
        data_dir =  os.path.join( todays_dir, RepositoryId )
        if os.path.exists( data_dir ):
            dir_index = self.get_next_numeric_directory_index(data_dir)
        else:
            dir_index = 1
        working_dir = os.path.join(data_dir, str(dir_index) )
#            os.makedirs(working_dir)
        return working_dir


    def get_new_entry_at_data_location(self, DataDirectory, RepositoryId):
        date_stamp = '%4d-%02d-%02d' % time.localtime()[:3]
        todays_dir = os.path.join( DataDirectory, date_stamp )
        data_dir =  os.path.join( todays_dir, RepositoryId )
        if os.path.exists( data_dir ):
            dir_index = self.get_next_numeric_directory_index(data_dir)
        else:
            dir_index = 1
        working_dir = os.path.join(data_dir, str(dir_index) )
        return working_dir



    #TODO Include check of if e.g. the network disk is alive
    def is_existing_remote_repository(self):
        return True




class TempFiles:
    """ The TempFilesLocation class is intended as a handler for the location of tempfiles.
        Better performance is achived if the temp-files are placed on a local high speed
        solid-state disk (SSD). It is also better to not have the tempfiles on the same
        physical disk as the raw data as this slows down the loading of data into the temp-file."""

    def __init__(self, NameOfSoftware):
        self.pure_software_name = os.path.basename(os.path.splitext(NameOfSoftware)[0])
        self.CurrentLocationFile = self.pure_software_name+'_TempFilesLocation.txt'
        if platform.system() == 'Darwin':
            self.os_settings_file_location = os.path.join(os.path.expanduser('~'),'.'+ self.pure_software_name )
        if platform.system() == 'Linux':
            self.os_settings_file_location = os.path.join(os.path.expanduser('~'),'.'+ self.pure_software_name )
        elif platform.system() == 'Windows':
            self.os_settings_file_location = os.path.join('C:\\ProgramData', self.pure_software_name)
        self.settings_file_path =  os.path.join(self.os_settings_file_location, self.CurrentLocationFile)
#        print( 'Tempfiles self.settings_file_path ', self.settings_file_path)


    def get_location(self):
        if not os.path.isfile(self.settings_file_path):
            return None
        else:
            settings_file = open(self.settings_file_path, 'r')
            TempFiles_directory = settings_file.readline()
            settings_file.close()
            if not os.path.exists(TempFiles_directory):
                return None
            else:
                return TempFiles_directory


    def set_location(self, location_of_tempfiles):
        make_dir_if_absent(os.path.dirname(self.settings_file_path))
        settings_file = open(self.settings_file_path, 'w')
        TempFiles_directory = os.path.join(location_of_tempfiles, self.pure_software_name)
        settings_file.write(TempFiles_directory)
        make_dir_if_absent(TempFiles_directory)
        settings_file.close()


    def delete_old_files(self, hours_to_expiry=8, file_extension='.tmp' ):
        FileLocation = self.get_location()
        files0 = os.listdir(FileLocation)
        for entry in sorted(files0):
            t_entry = os.path.getmtime(os.path.join(FileLocation, entry))
            modification_time = datetime.datetime.fromtimestamp(t_entry)
            is_old = (datetime.datetime.today() - modification_time) > datetime.timedelta(hours=hours_to_expiry)
            if (os.path.splitext(entry)[1] == file_extension) and is_old:
                print( 'Removing temp file: ', entry )
                # print( modification_date(os.path.join(FileLocation, entry)), entry
                # print( os.path.splitext(entry)[0] +'.txt'
                os.remove(os.path.join(FileLocation, entry))
